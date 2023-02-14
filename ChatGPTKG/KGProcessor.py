import os
import openpyxl


def process_vertexes(file_name, output_file):
    with open(file_name, 'r') as f:
        lines = f.readlines()
    data = [line.strip().replace("结局：", "").replace("：", ",").replace("，",",").split(',') for line in lines if line.strip()]

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Vertexes"

    worksheet['A1'] = "实体id"
    worksheet['B1'] = "实体名"
    worksheet['C1'] = "实体标签"
    worksheet['D1'] = "引导词"

    property_names = ["性别", "职业", "性格", "特长", "著名之处", "结局"]
    for i, property_name in enumerate(property_names, start=1):
        worksheet.cell(row=1, column=4 + 2 * i - 1, value=f"属性名_{i}")
        worksheet.cell(row=1, column=4 + 2 * i, value=f"属性值_{i}")

    names = set()
    row_index = 2
    for i, row in enumerate(data):
        if len(row) < 3:
            continue
        name = row[0]

        name_sec = name.split(" ")
        if len(name_sec) > 2:
            name = name_sec[0] + " " + name_sec[-1]

        if name in names:
            continue
        names.add(name)

        worksheet.cell(row=row_index, column=1, value=name)
        worksheet.cell(row=row_index, column=2, value=name)
        worksheet.cell(row=row_index, column=3, value=row[1])
        worksheet.cell(row=row_index, column=4, value='')
        property_index = 0
        col_index = 5
        for cell_value in row[2:]:
            prompt_name = property_names[property_index]
            worksheet.cell(row=row_index, column=col_index, value=prompt_name)
            worksheet.cell(row=row_index, column=col_index + 1, value=cell_value)
            col_index += 2
            property_index += 1
        row_index += 1
    workbook.save(output_file)
    return names


def process_edges(names, file_path, xlsx_path):
    if not os.path.exists(file_path):
        print(f"{file_path} not exists.")
        return

    if os.path.exists(xlsx_path):
        workbook = openpyxl.load_workbook(xlsx_path)
    else:
        workbook = openpyxl.Workbook()

    worksheet = workbook.create_sheet("Edges")
    worksheet.cell(row=1, column=1).value = "关系类型"
    worksheet.cell(row=1, column=2).value = "源实体id"
    worksheet.cell(row=1, column=3).value = "目标实体id"

    fixed_relations = {"丈夫":"夫妻","妻子":"夫妻", "伯父":"亲戚","父子":"父亲","母子":"母亲","导师":"教师","师徒":"教师"}
    reverse_relations = {"朋友":"朋友", "同学":"同学", "敌人":"敌人", "亲戚":"亲戚", "夫妻":"夫妻", "教师":"学生", "学生":"教师","父亲":"子女","母亲":"子女"}

    pairs = set()
    with open(file_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()
        i = 2
        for line in lines:
            line = line.strip()
            if line == '':
                continue
            line = line.replace("，", ",")
            elements = line.split(",")
            relationship = elements[0]
            source = elements[1].strip()
            targets = elements[2].strip().split("、")
            for target in targets:
                if relationship in fixed_relations.keys():
                    relationship = fixed_relations[relationship]
                # 因为通过ChatGPT抽取的关系与SmartKG预设的关系逻辑正好相反，因此需要反向处理关系的方向
                i, pairs = process_relation(i, names, relationship,  target, source, worksheet, pairs)
                if relationship in reverse_relations.keys():
                    reverse_relationship = reverse_relations[relationship]
                    i, pairs = process_relation(i, names, reverse_relationship,  source, target, worksheet,pairs)

    workbook.save(xlsx_path)

    norelationshipt_name = set()
    
    for name in names:
        find = False
        for pair in pairs:
            if pair.find(name) != -1:
                find = True
                continue
        if not find:
            norelationshipt_name.add(name)


    print("Following name are not listed in any relationship:\n")
    print(norelationshipt_name)
    return


def process_relation(row_index, names, relationship, source, target, worksheet, pairs):
    if source in names and target in names:
        pair = relationship + ":" + source + "_" + target
        if pair in pairs:
            return row_index, pairs

        pairs.add(pair)
        worksheet.cell(row=row_index, column=1, value=relationship)
        worksheet.cell(row=row_index, column=2, value=source)
        worksheet.cell(row=row_index, column=3, value=target)
        row_index += 1
    return row_index, pairs


output_file = 'data\\HarryPotter_KG.xlsx'
names = process_vertexes('data\\vertexes.txt', output_file)
process_edges(names, "data\\edges.txt", output_file)